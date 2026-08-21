#!/usr/bin/env python3
"""Extract cell values + formatting (color, strike, fill, comment) from an xlsx file.

Usage: extract-cells.py <xlsx_path>
Prints JSON: {"<sheet name>": [{"cell": "A1", "value": ..., "font_color": ..., "strike": ..., "fill_color": ..., "comment": ...}, ...]}
"""
import datetime
import json
import sys

import openpyxl


def _json_default(obj):
    if isinstance(obj, (datetime.datetime, datetime.date, datetime.time)):
        return obj.isoformat()
    raise TypeError(f"Object of type {obj.__class__.__name__} is not JSON serializable")


def _rgb(color_obj):
    if color_obj is None:
        return None
    rgb = getattr(color_obj, "rgb", None)
    return rgb if isinstance(rgb, str) else None


def extract(path: str) -> dict:
    # Two loads on purpose: data_only=True resolves formulas to their cached
    # value, but once resolved a cell's data_type no longer says "formula" —
    # there is no way to tell "formula with no cached value" apart from "truly
    # empty cell" from this load alone. The data_only=False load keeps the raw
    # formula string so that distinction can be made explicit instead of the
    # cell silently vanishing (a real bug this replaced: an unopened/
    # unrecalculated file — e.g. written by a script, never opened in Excel/
    # LibreOffice — would drop every formula cell with zero warning).
    wb_values = openpyxl.load_workbook(path, data_only=True)
    wb_formulas = openpyxl.load_workbook(path, data_only=False)
    sheets = {}
    for ws in wb_values.worksheets:
        ws_formulas = wb_formulas[ws.title]
        cells = []
        for row in ws.iter_rows():
            for cell in row:
                raw = ws_formulas[cell.coordinate].value
                is_formula = isinstance(raw, str) and raw.startswith("=")
                if cell.value is None:
                    if not is_formula:
                        continue
                    # Formula cell, no cached value available.
                    cells.append({
                        "cell": cell.coordinate,
                        "value": None,
                        "formula": raw,
                        "warning": "formula has no cached value — source file was "
                                   "likely never opened/saved in Excel or LibreOffice; "
                                   "re-save it there to populate a cached value",
                    })
                    continue
                font = cell.font
                fill = cell.fill
                cells.append({
                    "cell": cell.coordinate,
                    "value": cell.value,
                    "formula": raw if is_formula else None,
                    "font_color": _rgb(font.color) if font else None,
                    "strike": bool(font.strike) if font else False,
                    "fill_color": _rgb(fill.fgColor) if fill else None,
                    "comment": cell.comment.text.strip() if cell.comment else None,
                })
        sheets[ws.title] = cells
    return sheets


def main() -> None:
    if len(sys.argv) != 2:
        print("usage: extract-cells.py <xlsx_path>", file=sys.stderr)
        sys.exit(2)
    print(json.dumps(extract(sys.argv[1]), ensure_ascii=False, indent=2, default=_json_default))


if __name__ == "__main__":
    main()
