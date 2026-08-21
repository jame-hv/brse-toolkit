import datetime
import json
import subprocess
import sys
import pathlib

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.comments import Comment

SCRIPT = pathlib.Path(__file__).resolve().parents[1] / "extract-cells.py"


def make_fixture(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ScreenA"
    ws["A1"] = "必須項目"
    ws["A1"].font = Font(color="FFFF0000", strike=True)
    ws["A1"].fill = PatternFill(fill_type="solid", fgColor="FFFFFF00")
    ws["A1"].comment = Comment("cần xác nhận lại", "brse")
    ws["B1"] = "登録"
    wb.save(path)


def run(xlsx_path):
    result = subprocess.run(
        [sys.executable, str(SCRIPT), str(xlsx_path)],
        capture_output=True, text=True, encoding="utf-8",
    )
    return json.loads(result.stdout), result.returncode


def test_extracts_value_format_and_comment(tmp_path):
    xlsx = tmp_path / "fixture.xlsx"
    make_fixture(xlsx)
    out, code = run(xlsx)
    assert code == 0
    cells = {c["cell"]: c for c in out["ScreenA"]}
    assert cells["A1"]["value"] == "必須項目"
    assert cells["A1"]["strike"] is True
    assert cells["A1"]["comment"] == "cần xác nhận lại"
    assert cells["B1"]["value"] == "登録"
    assert cells["B1"]["strike"] is False


def test_empty_cells_are_skipped(tmp_path):
    xlsx = tmp_path / "fixture.xlsx"
    make_fixture(xlsx)
    out, _ = run(xlsx)
    coords = {c["cell"] for c in out["ScreenA"]}
    assert "C1" not in coords


def test_date_cells_are_serialized(tmp_path):
    xlsx = tmp_path / "fixture_date.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "History"
    ws["A1"] = datetime.datetime(2026, 8, 11)
    wb.save(xlsx)
    out, code = run(xlsx)
    assert code == 0
    cells = {c["cell"]: c for c in out["History"]}
    assert cells["A1"]["value"] == "2026-08-11T00:00:00"


def test_formula_with_no_cached_value_is_reported_not_dropped(tmp_path):
    """Regression: a formula cell in a file that was never opened/recalculated
    by Excel/LibreOffice has no cached value under data_only=True — it used to
    be silently skipped, indistinguishable from a genuinely empty cell."""
    xlsx = tmp_path / "fixture_formula.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Calc"
    ws["A1"] = "=1+1"  # openpyxl never recalculates formulas itself
    wb.save(xlsx)
    out, code = run(xlsx)
    assert code == 0
    cells = {c["cell"]: c for c in out["Calc"]}
    assert cells["A1"]["value"] is None
    assert cells["A1"]["formula"] == "=1+1"
    assert "warning" in cells["A1"]


def test_plain_value_cell_reports_formula_as_none(tmp_path):
    """A non-formula cell is unaffected by the two-pass (values + raw) load —
    `formula` stays None rather than picking up noise from the second load.
    (Faking a *cached* formula value needs a real recalculation engine, which
    openpyxl doesn't have — the no-cache path above is the one that matters:
    it's the one that used to silently lose data.)"""
    xlsx = tmp_path / "fixture_plain.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Calc"
    ws["A2"] = 4
    wb.save(xlsx)
    out, code = run(xlsx)
    cells = {c["cell"]: c for c in out["Calc"]}
    assert cells["A2"]["value"] == 4
    assert cells["A2"]["formula"] is None


def test_merged_cell_only_top_left_has_a_value(tmp_path):
    """Documents existing openpyxl/xlsx behavior (not a bug to fix here): a
    merged range's value lives only on the top-left cell, the rest read as
    empty — callers must not assume every cell in a merged header has data."""
    xlsx = tmp_path / "fixture_merge.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Header"
    ws["A1"] = "merged label"
    ws.merge_cells("A1:C1")
    wb.save(xlsx)
    out, code = run(xlsx)
    coords = {c["cell"] for c in out["Header"]}
    assert coords == {"A1"}
